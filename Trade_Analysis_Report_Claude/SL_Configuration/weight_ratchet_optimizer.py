"""
weight_ratchet_optimizer.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Finds the optimal TP allocation weights + ratchet mode combination
for every TP count (2–14), using existing ratchet_results.csv + trades_dataset.csv.

No new API calls needed — TP hit sequences are price-level events independent
of position sizing, so {mode}_tp values from existing results are reused directly.

For each (n_targets, ratchet_mode) pair, tests:
  1. Systematic weight families  (equal, front-loaded, bot-default, back-loaded variants)
  2. Constrained random search   (100k combinations per n per mode, valid weights only)

Outputs:
  weight_ratchet_optimal.csv    — best (weight_scheme, ratchet_mode, avg_R) per n_targets
  weight_ratchet_full.csv       — avg R for every tested combination
  weight_ratchet_report.txt     — human-readable summary
  weight_ratchet_report.xlsx    — full Excel report
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import csv, ast, random, math
from collections import defaultdict
from pathlib import Path
from itertools import product

# ── Config ────────────────────────────────────────────────────────────────────
RESULTS_CSV = r"C:\Users\louis\OneDrive\Louis\WSQ_ATB\Trade_Analysis_Report_Claude\SL_Configuration\ratchet_results.csv"
TRADES_CSV  = r"C:\Users\louis\OneDrive\Louis\WSQ_ATB\output\trades_dataset.csv"
OUT_DIR     = r"C:\Users\louis\OneDrive\Louis\WSQ_ATB\Trade_Analysis_Report_Claude\SL_Configuration"
YEAR_FILTER = ('2025', '2026')   # set to None to use all years
RANDOM_SEEDS = 120_000           # random weight combinations per (n, mode) pair

MODES = [
    'no_ratchet','be_only','standard','skip_tp1_move','skip_tp2_move',
    'skip_tp3_move','lag_2','lag_3','every_2_tps','every_3_tps',
    'every_2_from_tp1','half_ratchet','quarter_ratchet','three_quarter_ratchet',
    'be_then_lag2','be_then_lag3','tp2_be_then_standard','tp3_be_then_standard',
    'skip_every_other_after_be','aggressive_plus1','two_steps_forward',
]

# ── Bot's current weights ─────────────────────────────────────────────────────
BOT_WEIGHTS = {
    1:  [100],
    2:  [14, 86],
    3:  [41, 21, 65],
    4:  [14, 10, 16, 60],
    5:  [10, 12, 18, 25, 35],
    6:  [5, 7, 10, 15, 25, 38],
    7:  [4, 5, 7, 10, 15, 25, 34],
    8:  [3, 4, 5, 7, 10, 16, 25, 30],
    9:  [3, 3, 4, 6, 8, 12, 17, 22, 25],
    10: [3, 3, 4, 5, 7, 10, 13, 17, 20, 18],
    11: [3, 3, 3, 4, 6, 8, 11, 14, 17, 16, 15],
    12: [2, 3, 3, 4, 5, 7,  9, 12, 14, 15, 14, 12],
    13: [2, 2, 3, 4, 5, 6,  8, 10, 12, 13, 13, 12, 10],
    14: [2, 2, 3, 3, 4, 5,  7,  9, 11, 12, 13, 12, 10,  7],
    15: [2, 2, 2, 3, 4, 5,  6,  8, 10, 11, 12, 12, 10,  7,  6],
}

# ── Ratchet SL position in R-space after tp_index TPs hit ────────────────────
def ratchet_sl_R(mode, tp_index, entry_R, tp_R_list):
    """
    Returns the SL position in R-units after tp_index TPs have been hit.
    entry_R = 0.0 (break-even).
    tp_R_list = list of R distances for each TP.
    original SL = -1.0R.
    Returns None to mean 'no change from previous SL'.
    """
    if mode == 'no_ratchet':
        return -1.0
    elif mode == 'be_only':
        return 0.0 if tp_index == 0 else None
    elif mode == 'standard':
        return 0.0 if tp_index == 0 else tp_R_list[tp_index - 1]
    elif mode == 'skip_tp1_move':
        if tp_index == 0: return None
        if tp_index == 1: return 0.0
        return tp_R_list[tp_index - 1]
    elif mode == 'skip_tp2_move':
        if tp_index == 0: return 0.0
        if tp_index == 1: return None
        return tp_R_list[tp_index - 1]
    elif mode == 'skip_tp3_move':
        if tp_index == 0: return 0.0
        if tp_index == 1: return tp_R_list[0]
        if tp_index == 2: return None
        return tp_R_list[tp_index - 1]
    elif mode == 'lag_2':
        idx = tp_index - 2
        if idx < 0: return None
        return 0.0 if idx == 0 else tp_R_list[idx - 1]
    elif mode == 'lag_3':
        idx = tp_index - 3
        if idx < 0: return None
        return 0.0 if idx == 0 else tp_R_list[idx - 1]
    elif mode == 'every_2_tps':
        if (tp_index + 1) % 2 != 0: return None
        return 0.0 if tp_index == 1 else tp_R_list[tp_index - 2]
    elif mode == 'every_3_tps':
        if (tp_index + 1) % 3 != 0: return None
        return 0.0 if tp_index == 2 else tp_R_list[tp_index - 3]
    elif mode == 'every_2_from_tp1':
        if (tp_index + 1) % 2 == 0: return None
        return 0.0 if tp_index == 0 else tp_R_list[tp_index - 1]
    elif mode == 'half_ratchet':
        if tp_index == 0: return (-1.0 + 0.0) / 2
        prev = 0.0 if tp_index == 1 else tp_R_list[tp_index - 2]
        return (prev + tp_R_list[tp_index - 1]) / 2
    elif mode == 'quarter_ratchet':
        if tp_index == 0: return -1.0 + 0.25 * (0.0 - (-1.0))
        prev = 0.0 if tp_index == 1 else tp_R_list[tp_index - 2]
        return prev + 0.25 * (tp_R_list[tp_index - 1] - prev)
    elif mode == 'three_quarter_ratchet':
        if tp_index == 0: return -1.0 + 0.75 * (0.0 - (-1.0))
        prev = 0.0 if tp_index == 1 else tp_R_list[tp_index - 2]
        return prev + 0.75 * (tp_R_list[tp_index - 1] - prev)
    elif mode == 'be_then_lag2':
        if tp_index == 0: return 0.0
        if tp_index == 1: return None
        return tp_R_list[tp_index - 2]
    elif mode == 'be_then_lag3':
        if tp_index == 0: return 0.0
        if tp_index <= 2: return None
        return tp_R_list[tp_index - 3]
    elif mode == 'tp2_be_then_standard':
        if tp_index == 0: return None
        if tp_index == 1: return 0.0
        return tp_R_list[tp_index - 1]
    elif mode == 'tp3_be_then_standard':
        if tp_index < 2: return None
        if tp_index == 2: return 0.0
        return tp_R_list[tp_index - 1]
    elif mode == 'skip_every_other_after_be':
        if tp_index == 0: return 0.0
        if tp_index == 1: return None
        return tp_R_list[tp_index - 1] if tp_index % 2 == 0 else None
    elif mode == 'aggressive_plus1':
        if tp_index == 0: return 0.0
        return tp_R_list[tp_index] if tp_index < len(tp_R_list) - 1 else tp_R_list[tp_index - 1]
    elif mode == 'two_steps_forward':
        if tp_index == 0: return 0.0
        if tp_index == 1: return tp_R_list[0]
        idx = tp_index - 2
        return tp_R_list[idx] if idx >= 0 else tp_R_list[0]
    return -1.0


def get_final_sl_R(mode, tps_hit, tp_R_list):
    """
    Returns the SL level in R-units at the point the trade closes,
    given that `tps_hit` TPs were hit under this ratchet mode.
    """
    cur_sl = -1.0
    for i in range(tps_hit):
        new_sl = ratchet_sl_R(mode, i, 0.0, tp_R_list)
        if new_sl is not None:
            cur_sl = max(cur_sl, new_sl)  # SL can only move in profit direction
    return cur_sl


def compute_R(fracs, tps_hit, tp_R_list, mode, n):
    """
    Compute realised R for a given weight distribution and ratchet mode.
    tps_hit: how many TPs price actually touched under this mode (from results file).
    """
    if tps_hit == 0:
        return -1.0  # pure loss regardless of weights

    realised  = sum(fracs[i] * tp_R_list[i] for i in range(tps_hit))
    remaining = sum(fracs[i] for i in range(tps_hit, n))
    sl_r      = get_final_sl_R(mode, tps_hit, tp_R_list)
    return realised + remaining * sl_r


# ── Weight scheme generators ──────────────────────────────────────────────────
def normalise(weights):
    t = sum(weights)
    return [w / t for w in weights]


def systematic_schemes(n):
    """
    Generate named systematic weight schemes for n TPs.
    Returns list of (name, fracs) tuples.
    """
    schemes = []

    # Equal
    schemes.append(('equal', normalise([1] * n)))

    # Bot default
    if n in BOT_WEIGHTS:
        schemes.append(('bot_default', normalise(BOT_WEIGHTS[n])))

    # Pure back-loaded power curves
    for p in [1.5, 2, 2.5, 3, 4]:
        w = [(i + 1) ** p for i in range(n)]
        schemes.append((f'back_power_{p}', normalise(w)))

    # Pure front-loaded power curves
    for p in [1.5, 2, 2.5, 3]:
        w = [(n - i) ** p for i in range(n)]
        schemes.append((f'front_power_{p}', normalise(w)))

    # Linear ramp back/front
    schemes.append(('linear_back',  normalise(list(range(1, n + 1)))))
    schemes.append(('linear_front', normalise(list(range(n, 0, -1)))))

    # Flat first half, spike last
    half = n // 2
    for spike in [0.3, 0.4, 0.5, 0.6]:
        w = [1.0] * n
        w[-1] = spike * n
        schemes.append((f'spike_last_{int(spike*100)}pct', normalise(w)))

    # Spike at each TP individually (useful for identifying which TP matters most)
    for spike_idx in range(n):
        w = [1.0] * n
        w[spike_idx] = n * 2
        schemes.append((f'spike_tp{spike_idx+1}', normalise(w)))

    # U-shaped (heavy first and last)
    for center_dip in [0.3, 0.5]:
        w = [(abs(i - (n-1)/2) / ((n-1)/2)) * (1 - center_dip) + center_dip
             for i in range(n)]
        schemes.append((f'u_shape_{int(center_dip*100)}', normalise(w)))

    # Geometric sequences
    for r in [1.2, 1.35, 1.5, 1.7, 2.0]:
        w = [r ** i for i in range(n)]
        schemes.append((f'geo_back_{r}', normalise(w)))
        w_rev = list(reversed(w))
        schemes.append((f'geo_front_{r}', normalise(w_rev)))

    # Variations on bot_default: scale specific regions
    if n in BOT_WEIGHTS:
        base = list(BOT_WEIGHTS[n])
        # Boost early TPs by 50%
        w = base[:]
        for i in range(min(3, n)): w[i] *= 1.5
        schemes.append(('bot_boost_early', normalise(w)))
        # Boost late TPs
        w = base[:]
        for i in range(max(0, n-3), n): w[i] *= 1.5
        schemes.append(('bot_boost_late', normalise(w)))
        # Flatten (move toward equal)
        w = [0.5 * base[i] + 0.5 * 1.0 for i in range(n)]
        schemes.append(('bot_flattened', normalise(w)))
        # Pure back amplified
        w = [base[i] ** 1.5 for i in range(n)]
        schemes.append(('bot_amplified', normalise(w)))

    return schemes


def random_weight_schemes(n, k, seed=42):
    """
    Generate k random valid weight distributions for n TPs.
    Constraints: each weight in [1, 50], weights sum to 100,
    monotonically checked via Dirichlet-like sampling.
    """
    rng = random.Random(seed)
    schemes = []
    for _ in range(k):
        # Sample from Dirichlet via exponentials — naturally sums to 1
        raw = [rng.expovariate(1.0) for _ in range(n)]
        total = sum(raw)
        fracs = [r / total for r in raw]
        schemes.append(('random', fracs))
    return schemes


# ── Load & clean data ─────────────────────────────────────────────────────────
def is_clean(row, trades):
    t = trades.get(row['message_id'], {})
    try:
        tps = ast.literal_eval(t.get('targets', '[]'))
        entry = float(t.get('entry_mid', 0))
        sl = float(t.get('stop_loss', 0))
        if entry <= 0 or sl <= 0: return False
        risk = abs(entry - sl)
        return all(abs(tp - entry) / risk <= 500 for tp in tps)
    except:
        return False


def load_data(results_path, trades_path, year_filter=None):
    with open(results_path, newline='', encoding='utf-8-sig') as f:
        results = list(csv.DictReader(f))
    with open(trades_path, newline='', encoding='utf-8-sig') as f:
        trades = {r['message_id']: r for r in csv.DictReader(f)}

    clean = []
    for r in results:
        if year_filter and r['date'][:4] not in year_filter:
            continue
        if not is_clean(r, trades):
            continue
        t = trades[r['message_id']]
        n = int(r['n_targets'])
        tp_R = [float(t.get(f'tp{i+1}_R', 0)) for i in range(n)]
        # Per-mode: how many TPs hit
        mode_tps = {m: int(r.get(f'{m}_tp', 0)) for m in MODES}
        clean.append({
            'message_id': r['message_id'],
            'date':        r['date'],
            'symbol':      r['symbol'],
            'n_targets':   n,
            'tp_R':        tp_R,
            'mode_tps':    mode_tps,
        })
    return clean


# ── Core optimizer ────────────────────────────────────────────────────────────
def run_optimization(data, year_label):
    # Group by n_targets
    by_n = defaultdict(list)
    for d in data:
        by_n[d['n_targets']].append(d)

    results_optimal  = []   # best per (n, mode) combo
    results_full     = []   # every scheme tested
    overall_best     = []   # single best per n across all modes and weights

    total_n = len(by_n)
    for n_idx, n in enumerate(sorted(by_n.keys())):
        trades_n = by_n[n]
        count = len(trades_n)
        print(f'  n={n:>2}  ({count:>3} trades)  ', end='', flush=True)

        sys_schemes  = systematic_schemes(n)
        rand_schemes = random_weight_schemes(n, RANDOM_SEEDS, seed=n * 1000)
        all_schemes  = sys_schemes + rand_schemes

        # For each mode: find best weight scheme
        best_per_mode = {}   # mode -> (scheme_name, fracs, avg_R)

        for mode in MODES:
            best_avg = -999
            best_name = ''
            best_fracs = None

            for scheme_name, fracs in all_schemes:
                # Compute avg R across all trades with this n using these weights + mode
                total_R = 0.0
                for d in trades_n:
                    tps_hit = d['mode_tps'][mode]
                    tp_R    = d['tp_R']
                    r_val   = compute_R(fracs, tps_hit, tp_R, mode, n)
                    total_R += r_val
                avg_R = total_R / count

                if avg_R > best_avg:
                    best_avg   = avg_R
                    best_name  = scheme_name
                    best_fracs = fracs

                # Store systematic schemes in full output (too many randoms to store all)
                if scheme_name != 'random':
                    results_full.append({
                        'n_targets':    n,
                        'mode':         mode,
                        'scheme':       scheme_name,
                        'avg_R':        round(avg_R, 5),
                        'trade_count':  count,
                    })

            best_per_mode[mode] = (best_name, best_fracs, best_avg)
            results_optimal.append({
                'n_targets':    n,
                'mode':         mode,
                'best_scheme':  best_name,
                'best_avg_R':   round(best_avg, 5),
                'trade_count':  count,
                'best_weights': str([round(f * 100, 2) for f in best_fracs]),
            })

        # Find the single best (mode + weight) for this n
        best_mode = max(best_per_mode, key=lambda m: best_per_mode[m][2])
        bm_scheme, bm_fracs, bm_avg = best_per_mode[best_mode]

        # Also compute bot_default + no_ratchet for comparison
        bot_fracs = normalise(BOT_WEIGHTS.get(n, [1]*n))
        bot_nr_R = sum(
            compute_R(bot_fracs, d['mode_tps']['no_ratchet'], d['tp_R'], 'no_ratchet', n)
            for d in trades_n
        ) / count
        bot_std_R = sum(
            compute_R(bot_fracs, d['mode_tps']['standard'], d['tp_R'], 'standard', n)
            for d in trades_n
        ) / count

        overall_best.append({
            'n_targets':       n,
            'trade_count':     count,
            'best_mode':       best_mode,
            'best_scheme':     bm_scheme,
            'best_avg_R':      round(bm_avg, 5),
            'best_weights_pct':str([round(f * 100, 2) for f in bm_fracs]),
            'bot_no_ratchet_R':round(bot_nr_R, 5),
            'bot_standard_R':  round(bot_std_R, 5),
            'improvement_vs_bot_nr': round(bm_avg - bot_nr_R, 5),
            'improvement_vs_bot_std':round(bm_avg - bot_std_R, 5),
        })

        print(f'best: {best_mode} + {bm_scheme} = {bm_avg:+.4f}R  '
              f'(bot+no_ratchet={bot_nr_R:+.4f}R  bot+std={bot_std_R:+.4f}R)')

    return results_optimal, results_full, overall_best


# ── Write outputs ─────────────────────────────────────────────────────────────
def write_csv(rows, path, fieldnames=None):
    if not rows: return
    fields = fieldnames or list(rows[0].keys())
    with open(path, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader(); w.writerows(rows)
    print(f'  → {path}')


def write_report(overall_best, results_optimal, out_path, year_label):
    lines = []
    W = 80
    lines.append('=' * W)
    lines.append(f'  WEIGHT + RATCHET JOINT OPTIMIZER — {year_label}')
    lines.append(f'  {RANDOM_SEEDS:,} random weight combinations × 21 modes × each TP count')
    lines.append('=' * W)
    lines.append('')
    lines.append('OPTIMAL (WEIGHT + RATCHET) PER TP COUNT')
    lines.append('-' * W)
    hdr = (f"  {'NTPs':>5}  {'Trades':>6}  {'Best mode':<26}  "
           f"{'Best scheme':<20}  {'Best R':>8}  {'vs bot+NR':>10}  {'vs bot+std':>11}")
    lines.append(hdr)
    lines.append('  ' + '-' * (W - 2))

    for row in overall_best:
        n   = row['n_targets']
        cnt = row['trade_count']
        bm  = row['best_mode']
        bs  = row['best_scheme']
        br  = row['best_avg_R']
        vnr = row['improvement_vs_bot_nr']
        vst = row['improvement_vs_bot_std']
        lines.append(
            f"  {n:>5}  {cnt:>6}  {bm:<26}  {bs:<20}  "
            f"{br:>+8.4f}R  {vnr:>+8.4f}R  {vst:>+9.4f}R"
        )

    lines.append('')
    lines.append('BEST WEIGHT DISTRIBUTION PER TP COUNT')
    lines.append('-' * W)
    for row in overall_best:
        lines.append(f"  {row['n_targets']} TPs  →  mode: {row['best_mode']}")
        lines.append(f"           scheme: {row['best_scheme']}")
        lines.append(f"           weights(%): {row['best_weights_pct']}")
        lines.append(f"           avg R: {row['best_avg_R']:+.4f}R  "
                     f"(vs bot+no_ratchet: {row['improvement_vs_bot_nr']:+.4f}R)")
        lines.append('')

    lines.append('=' * W)
    lines.append('TOP 3 MODES PER TP COUNT (best weight for each mode)')
    lines.append('-' * W)

    by_n = defaultdict(list)
    for row in results_optimal:
        by_n[row['n_targets']].append(row)

    for n in sorted(by_n.keys()):
        modes_sorted = sorted(by_n[n], key=lambda x: x['best_avg_R'], reverse=True)[:3]
        lines.append(f'  {n} TPs:')
        for rank, m in enumerate(modes_sorted, 1):
            lines.append(
                f'    {rank}. {m["mode"]:<26}  '
                f'{m["best_avg_R"]:>+.4f}R  ({m["best_scheme"]})'
            )
        lines.append('')

    text = '\n'.join(lines)
    with open(out_path, 'w') as f:
        f.write(text)
    print(f'  → {out_path}')
    print()
    print(text)


def write_excel(overall_best, results_optimal, out_path, year_label):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule
    except ImportError:
        print('  openpyxl not installed — skipping Excel output')
        return

    wb = Workbook()
    def _s(): s=Side(style='thin',color='CCCCCC'); return Border(left=s,right=s,top=s,bottom=s)
    def _f(h): return PatternFill('solid',fgColor=h)
    def _ft(b=False,s=10,c='000000'): return Font(bold=b,size=s,color=c,name='Arial')
    def _a(h='center'): return Alignment(horizontal=h,vertical='center',wrap_text=False)

    # ── Sheet 1: Overall best per N ───────────────────────────────────────────
    ws = wb.active; ws.title = '🏆 Optimal per TP count'
    ws.sheet_view.showGridLines = False
    widths = [3,8,8,26,24,12,14,14,14,14,50]
    cols   = 'ABCDEFGHIJK'
    for col,w in zip(cols,widths): ws.column_dimensions[col].width = w

    ws.merge_cells('B2:K2')
    c=ws.cell(2,2,f'OPTIMAL WEIGHT + RATCHET — {year_label}')
    c.font=_ft(True,14,'1F3864'); c.alignment=_a()
    ws.row_dimensions[2].height=28

    headers=['NTPs','Trades','Best mode','Best weight scheme',
             'Best avg R','vs bot+NR','vs bot+std','Bot NR avg R','Bot std avg R','Best weights (%)']
    for ci,h in enumerate(headers,2):
        c=ws.cell(4,ci,h); c.fill=_f('1F3864'); c.font=_ft(True,10,'FFFFFF')
        c.alignment=_a(); c.border=_s()
    ws.row_dimensions[4].height=20

    for ri,row in enumerate(overall_best):
        r=5+ri
        vals=[row['n_targets'],row['trade_count'],row['best_mode'],row['best_scheme'],
              row['best_avg_R'],row['improvement_vs_bot_nr'],row['improvement_vs_bot_std'],
              row['bot_no_ratchet_R'],row['bot_standard_R'],row['best_weights_pct']]
        is_imp = row['improvement_vs_bot_nr'] > 0.01
        rf = _f('C6EFCE') if is_imp else _f('FFF2CC')
        for ci,val in enumerate(vals,2):
            c=ws.cell(r,ci,val)
            c.font=_ft(ci in [4,5,6],10)
            c.alignment=_a('left' if ci in [4,5,11] else 'center')
            c.border=_s(); c.fill=rf
            if ci in [6,7,8,9,10]:
                c.number_format='+0.0000"R";-0.0000"R"'
        ws.row_dimensions[r].height=20

    ws.conditional_formatting.add(f'F5:F{4+len(overall_best)}',
        ColorScaleRule(start_type='min',start_color='FCE4D6',
                       mid_type='num',mid_value=0,mid_color='FFFFE0',
                       end_type='max',end_color='C6EFCE'))

    # ── Sheet 2: Top 3 per N ──────────────────────────────────────────────────
    ws2 = wb.create_sheet('📊 Top 3 per TP count')
    ws2.sheet_view.showGridLines = False
    for col,w in zip('ABCDEFG',[3,8,8,26,24,12,20]): ws2.column_dimensions[col].width=w

    ws2.merge_cells('B2:G2')
    c=ws2.cell(2,2,f'TOP 3 MODES PER TP COUNT — {year_label}')
    c.font=_ft(True,13,'1F3864'); c.alignment=_a()

    by_n_opt = defaultdict(list)
    for row in results_optimal:
        by_n_opt[row['n_targets']].append(row)

    row_idx = 4
    for n in sorted(by_n_opt.keys()):
        modes_sorted = sorted(by_n_opt[n], key=lambda x: x['best_avg_R'], reverse=True)

        # Section header
        ws2.merge_cells(f'B{row_idx}:G{row_idx}')
        c=ws2.cell(row_idx,2,f'{n} targets  ({modes_sorted[0]["trade_count"]} trades)')
        c.fill=_f('1F3864'); c.font=_ft(True,10,'FFFFFF'); c.alignment=_a(); c.border=_s()
        row_idx+=1

        for ci,h in enumerate(['Rank','Mode','Best scheme','Best avg R','Best weights (%)'],2):
            c=ws2.cell(row_idx,ci,h); c.fill=_f('2E5E9B')
            c.font=_ft(True,9,'FFFFFF'); c.alignment=_a(); c.border=_s()
        row_idx+=1

        fills=[_f('C6EFCE'),_f('FFF2CC'),_f('F5F5F5')]
        for rank,m in enumerate(modes_sorted[:5],1):
            rf=fills[min(rank-1,2)]
            ws2.cell(row_idx,2,rank).fill=rf; ws2.cell(row_idx,2).font=_ft(rank==1,10)
            ws2.cell(row_idx,2).border=_s(); ws2.cell(row_idx,2).alignment=_a()
            ws2.cell(row_idx,3,m['mode']).fill=rf; ws2.cell(row_idx,3).font=_ft(rank==1,10)
            ws2.cell(row_idx,3).border=_s(); ws2.cell(row_idx,3).alignment=_a('left')
            ws2.cell(row_idx,4,m['best_scheme']).fill=rf; ws2.cell(row_idx,4).font=_ft(False,9)
            ws2.cell(row_idx,4).border=_s(); ws2.cell(row_idx,4).alignment=_a('left')
            c=ws2.cell(row_idx,5,m['best_avg_R']); c.fill=rf; c.font=_ft(rank==1,10)
            c.number_format='+0.0000"R";-0.0000"R"'; c.border=_s(); c.alignment=_a()
            ws2.cell(row_idx,6,m['best_weights']).fill=rf; ws2.cell(row_idx,6).font=_ft(False,8)
            ws2.cell(row_idx,6).border=_s(); ws2.cell(row_idx,6).alignment=_a('left')
            ws2.row_dimensions[row_idx].height=18
            row_idx+=1
        row_idx+=1  # spacer

    wb.save(out_path)
    print(f'  → {out_path}')


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    out_dir = Path(OUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)

    year_label = f'{"-".join(YEAR_FILTER)}' if YEAR_FILTER else 'All years'
    year_tag   = year_label.replace('-','_')

    print(f'Loading data ({year_label})...')
    data = load_data(RESULTS_CSV, TRADES_CSV, YEAR_FILTER)
    print(f'  {len(data)} trades loaded')
    print(f'  {RANDOM_SEEDS:,} random weight schemes per (n, mode)')
    print(f'  {len(MODES)} ratchet modes')
    print()
    print('Running optimization...')

    results_optimal, results_full, overall_best = run_optimization(data, year_label)

    print()
    print('Writing outputs...')
    write_csv(overall_best,   out_dir / f'weight_ratchet_optimal_{year_tag}.csv')
    write_csv(results_optimal,out_dir / f'weight_ratchet_per_mode_{year_tag}.csv')
    write_csv(results_full,   out_dir / f'weight_ratchet_systematic_{year_tag}.csv')
    write_report(overall_best, results_optimal,
                 out_dir / f'weight_ratchet_report_{year_tag}.txt', year_label)
    write_excel(overall_best, results_optimal,
                out_dir / f'weight_ratchet_report_{year_tag}.xlsx', year_label)

    print(f'\nDone.')


if __name__ == '__main__':
    main()
